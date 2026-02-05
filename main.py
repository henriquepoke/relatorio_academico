from openpyxl import Workbook, load_workbook

arquivo = load_workbook('alunos.xlsx')
planilha = arquivo['Alunos']

# Nome, Curso, Idade, Nota final, Data de Matrícula

maior = 0
soma_notas = 0
total = 0
nome_maior = ''

planilha_aprovados = Workbook()
planilha_aprovados.active.Title = 'Aprovados'
planilha_aprovados.active.append(['Nome', 'Curso', 'Idade', 'Nota final', 'Data de Matrícula'])

planilha_reprovados = Workbook()
planilha_reprovados.active.Title = 'Reprovados'
planilha_reprovados.active.append(['Nome', 'Curso', 'Idade', 'Nota final', 'Data de Matrícula'])

for linha in planilha.iter_rows(min_row=2, values_only=True):
    nome, curso, idade, nota_final, matricula = linha
    soma_notas += nota_final
    total = len(planilha['A']) - 1  # Subtrai 1 para não contar o cabeçalho

    if nota_final > maior:
        maior = nota_final
        nome_maior = nome

    if nota_final >= 7:
        planilha_aprovados.active.append([nome, curso, idade, nota_final, matricula.strftime('%d/%m/%Y')])
    else:
        planilha_reprovados.active.append([nome, curso, idade, nota_final, matricula.strftime('%d/%m/%Y')])

    media = soma_notas / total
### Exibir no terminal:
# Quantidade de aprovados e reprovados
# Nota média da turma
# Nome do aluno com a maior nota

planilha_aprovados.save('alunos_aprovados.xlsx')
planilha_reprovados.save('alunos_reprovados.xlsx')

print(f'Total de alunos: {len(planilha["A"]) - 1}')
print(f'Quantidade de aprovados: {len(planilha_aprovados.active["A"]) - 1}')
print(f'Quantidade de reprovados: {len(planilha_reprovados.active["A"]) - 1}')
print(f'Nota média da turma: {media:.2f}')
print(f'Aluno com a maior nota: {nome_maior} ({maior})')
